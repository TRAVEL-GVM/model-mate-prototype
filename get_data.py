import pandas as pd
import numpy as np
import streamlit as st
import yfinance as yf
import matplotlib.pyplot as plt
import seaborn as sns
import csv
import json
import urllib
import time
from pyjstat import pyjstat
from ecbdata import ecbdata

from datetime import datetime
from config import *
from io import BytesIO
import requests

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


def extract_data_from_bank_pt(series_id, variable_name):
    """
    Function to extract data from BPSTAT API.

    Arguments: series_id int
             variable_name str.
             If variable_name is None, variable_name is set to urls label.

    Returns:   pandas dataframe with Date and variable_name columns
    """

    BPSTAT_API_URL = "https://bpstat.bportugal.pt/data/v1"

    url = f"{BPSTAT_API_URL}/series/?lang=EN&series_ids={series_id}"
    series_info = requests.get(url).json()[0]

    domain_id = series_info["domain_ids"][0]
    dataset_id = series_info["dataset_id"]

    dataset_url = f"{BPSTAT_API_URL}/domains/{domain_id}/datasets/{dataset_id}/?lang=EN&series_ids={series_id}"
    dataset = pyjstat.Dataset.read(dataset_url)
    df = dataset.write('dataframe')

    df['Date'] = pd.to_datetime(df['Date'])
    if variable_name is None:
        variable_name = series_info['label']

    df = df.rename(columns={'value': variable_name})
    df = df[['Date', variable_name]]
    df['Date'] = pd.to_datetime(df['Date']).dt.date

    return df


def extract_data_from_ecb(key, start_date='2020-01'):
    """
    Function to extract data from ECB.

    Arguments: key str: URL key
               start_date str:  start date

    Returns:   pandas dataframe with TIME_PERIOD and OBS_VALUE columns
    """

    df = ecbdata.get_series(key,
                            start=start_date, detail='dataonly')

    df.TIME_PERIOD = pd.to_datetime(df.TIME_PERIOD)
    df = df[['TIME_PERIOD', 'OBS_VALUE']]

    return df


def extract_euribor_data_from_ecb(tenor, start_date):
    """
    Function to extract Euribor data.
    Extracted from ECB.
    Returns a dataframe with euribor data for a defined tenor from start_date until now.

    Params:
        - tenor (str): '3M' or '6M' or '1M' or '1Y'
        - startdate (str)

    Returns a dataframe with euribor data for the specified tenor from start_date until now.

    Usage example:  extract_euribor_data_from_ecb('1Y', '2020-01-01')
    """

    dict_keys = {
        '3M': 'FM.M.U2.EUR.RT.MM.EURIBOR6MD_.HSTA',
        '6M': 'FM.M.U2.EUR.RT.MM.EURIBOR3MD_.HSTA',
        '1M': 'FM.M.U2.EUR.RT.MM.EURIBOR1MD_.HSTA',
        '1Y': 'FM.M.U2.EUR.RT.MM.EURIBOR1YD_.HSTA'
    }

    df = extract_data_from_ecb(dict_keys[tenor], start_date)
    df.columns = ['Date', 'Euribor ' + tenor]

    return df


def extract_euribors(start_date):
    """
    Function to extract Euribor data for several tenors ('3M','6M','1M','1Y').
    Extracted from ECB.

    Params:
        - start_date (str)

    Returns a dataframe with euribor data for several tenor from start_date until now.

    Usage example:  extract_euribors('2020-01-01')
    """

    dict_keys = {
        '3M': 'FM.M.U2.EUR.RT.MM.EURIBOR6MD_.HSTA',
        '6M': 'FM.M.U2.EUR.RT.MM.EURIBOR3MD_.HSTA',
        '1M': 'FM.M.U2.EUR.RT.MM.EURIBOR1MD_.HSTA',
        '1Y': 'FM.M.U2.EUR.RT.MM.EURIBOR1YD_.HSTA'
    }

    df_aux = extract_data_from_ecb(dict_keys['1M'], start_date)
    df_aux.columns = ['Date', 'Euribor 1M']

    for tenor in ['3M', '6M', '1Y']:
        df_aux1 = extract_data_from_ecb(dict_keys[tenor], start_date)
        df_aux1.columns = ['Date', 'Euribor ' + tenor]

        df_aux = df_aux.merge(df_aux1, on='Date', how='left')
    df_aux['Date'] = pd.to_datetime(df_aux['Date']).dt.date

    return df_aux


def convert_df_to_excel(df):

    output = BytesIO()

    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max() + 2, len(col) + 2) # Calcula o tamanho máximo da coluna + padding
            worksheet.set_column(idx, idx, max_len)  # Define a largura de cada coluna

        # Formatar o cabeçalho (linha das colunas)
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'font_color': '#FFFFFF',
            'valign': 'center',
            'align': 'center',
            'fg_color': default_color1,
            'border': 1
        })

        cell_format = workbook.add_format({
            'align': 'center',  # Centraliza horizontalmente
            'valign': 'vcenter',  # Centraliza verticalmente
        })

        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)

        for row_num in range(1, len(df) + 1):
            worksheet.set_row(row_num, None, cell_format)

    output.seek(0)
    return output

############################################
def extract_data_from_bank_pt(series_id, variable_name):
    """
    Function to extract data from BPSTAT API.

    Arguments: series_id int
             variable_name str.
             If variable_name is None, variable_name is set to urls label.

    Returns:   pandas dataframe with Date and variable_name columns
    """

    BPSTAT_API_URL = "https://bpstat.bportugal.pt/data/v1"

    url = f"{BPSTAT_API_URL}/series/?lang=EN&series_ids={series_id}"
    series_info = requests.get(url).json()[0]

    domain_id = series_info["domain_ids"][0]
    dataset_id = series_info["dataset_id"]

    dataset_url = f"{BPSTAT_API_URL}/domains/{domain_id}/datasets/{dataset_id}/?lang=EN&series_ids={series_id}"
    dataset = pyjstat.Dataset.read(dataset_url)
    df = dataset.write('dataframe')

    df['Date'] = pd.to_datetime(df['Date'])
    if variable_name is None:
        variable_name = series_info['label']

    df = df.rename(columns={'value': variable_name})
    df = df[['Date', variable_name]]

    return df


def get_df_final():
    """Dados do bpstat sem multiindex"""
    ano_atual = datetime.datetime.now().year - 1

    df_final = pd.DataFrame()
    df_final['Date'] = pd.date_range(start='2006-12-31', end=f'{ano_atual}-12-31', freq='YE-DEC')

    for setor, series_ids in indicadores.items():
        df_setor = pd.DataFrame()
        for series_id in series_ids:
            df_extracted = extract_data_from_bank_pt(series_id, None)

            df_final = df_final.merge(df_extracted, on='Date', how='left')

    return df_final


def get_df_final_comh():
    """Dados do bpstat com multiindex"""
    ano_atual = datetime.datetime.now().year - 1

    df_final_comh = pd.DataFrame()
    df_final_comh['Date'] = pd.date_range(start='2006-12-31', end=f'{ano_atual}-12-31', freq='YE-DEC')

    colunas_nivel_1 = ['Date']
    colunas_nivel_2 = ['']

    for setor, series_ids in indicadores.items():
        df_setor = pd.DataFrame()
        for series_id in series_ids:
            df_extracted = extract_data_from_bank_pt(series_id, None)

            df_final_comh = df_final_comh.merge(df_extracted, on='Date', how='left')

            colunas_nivel_1.extend([setor] * (df_extracted.shape[1] - 1))
            colunas_nivel_2.extend(df_extracted.columns[1:])

    df_final_comh.columns = pd.MultiIndex.from_tuples(list(zip(colunas_nivel_1, colunas_nivel_2)))
    df_final_comh = df_final_comh.loc[:, ~df_final_comh.columns.duplicated()]

    df_final_comh = df_final_comh.set_index('Date')

    df_final_comh.columns = pd.MultiIndex.from_tuples(
        [(col[0], col[1].split('-')[1]) for col in df_final_comh.columns]
    )

    df_final_comh.index = df_final_comh.index.date

    return df_final_comh


def transform_data(d):
    new_data = {}
    for key, value in d.items():
        new_data[key] = {}
        for sub_key, url in value.items():
            if "Quarterly" in sub_key:
                periodicity = "Quarterly"
                new_key = sub_key.replace(", Quarterly", "")
            elif "Monthly" in sub_key:
                periodicity = "Monthly"
                new_key = sub_key.replace(", Monthly", "")
            elif "Annual" in sub_key:
                periodicity = "Annual"
                new_key = sub_key.replace(", Annual", "")
            else:
                periodicity = "Unknown"
                new_key = sub_key

            new_data[key][new_key] = {
                "periodicity": periodicity,
                "url": url
            }

    return new_data


def extract_and_combine_data(transformed_data, periodicity, start_date='2020-01'):
    combined_series = pd.DataFrame()

    for category, indicators in transformed_data.items():
        for indicator, details in indicators.items():
            if details['periodicity'] == periodicity:
                url = details['url'].split("datasets/")[1].split('/')[1]
                data = extract_data_from_ecb(url, start_date)
                if not data.empty:
                    data.rename(columns={'OBS_VALUE': indicator}, inplace=True)
                    if combined_series.empty:
                        combined_series = data
                    else:
                        combined_series = pd.merge(combined_series, data, on='TIME_PERIOD', how='outer')

    #combined_series['TIME_PERIOD'] = combined_series['TIME_PERIOD'].dt.date

    return combined_series

