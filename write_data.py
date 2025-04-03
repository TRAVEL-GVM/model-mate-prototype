from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

import pandas as pd
from config import *
from get_data import *

ano_atual = datetime.datetime.now().year - 1

df_final = pd.DataFrame()
df_final['Date'] = pd.date_range(start='2006-12-31', end=f'{ano_atual}-12-31', freq='A-DEC')

for setor, series_ids in indicadores.items():
    df_setor = pd.DataFrame()
    for series_id in series_ids:
        df_extracted = extract_data_from_bank_pt(series_id, None)

        df_final = df_final.merge(df_extracted, on='Date', how='left')

# df_final

df_final.to_excel("indicadores_bpstat_sem_header.xlsx", index=False)

print("Ficheiro Excel criado com sucesso.")

